from io import BytesIO
import openpyxl
from .base import BaseGenerator, BOLD, CENTER, LEFT


class IndividualProtocolGenerator(BaseGenerator):
    doc_type = "protocol_individual"
    title = "Индивидуальный протокол защиты"

    def generate(self) -> bytes:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for idx, student in enumerate(self.students, 1):
            safe_name = self._sheet_name(idx, student.full_name)
            ws = wb.create_sheet(title=safe_name)
            self._build_for_student(ws, student, idx)
        if not wb.sheetnames:
            wb.create_sheet(title="Пусто")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def build(self, ws):
        pass

    @staticmethod
    def _sheet_name(idx: int, full_name: str) -> str:
        forbidden = set('[]:*?/\\')
        clean = "".join(ch for ch in (full_name or "") if ch not in forbidden)
        name = f"{idx}_{clean}"
        return name[:31]

    def _build_for_student(self, ws, student, proto_no):
        s = self.session
        spec = s.specialty.name if s.specialty else ""
        chair = self.member_by_role("chairman")
        deputy = self.member_by_role("deputy")
        members = self.member_by_role("member")
        secretary = self.member_by_role("secretary")

        r = 1

        def put(text, bold=False, align=CENTER):
            nonlocal r
            cell = ws.cell(r, 1, text)
            if bold:
                cell.font = BOLD
            cell.alignment = align
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            r += 1

        put("ПРОТОКОЛ ЗАСЕДАНИЯ ГЭК (ЗАЩИТА)", bold=True)
        put(f"ПРОТОКОЛ № {proto_no}", bold=True)
        put(f"заседания Государственной аттестационной комиссии № {s.number}")
        put(f"от {self.fmt_date(s.event_date)}")
        put(f"по защите {s.work_type or 'дипломного проекта'}")
        put(f"по специальности {spec}")
        r += 1

        ws.cell(r, 1, "Присутствовали:")
        if chair:
            ws.cell(r, 3, "Председатель ГЭК:")
            ws.cell(r, 5, chair[0].short_name or chair[0].full_name)
        r += 1
        for m in deputy:
            ws.cell(r, 3, "Зам. Председателя")
            ws.cell(r, 5, m.short_name or m.full_name)
            r += 1
        for i, m in enumerate(members):
            ws.cell(r, 3, "Члены ГЭК:" if i == 0 else "")
            ws.cell(r, 5, m.short_name or m.full_name)
            r += 1
        for m in secretary:
            ws.cell(r, 3, "Секретарь ГЭК:")
            ws.cell(r, 5, m.short_name or m.full_name)
            r += 1
        r += 1

        ws.cell(r, 1, "Аттестуется студент(ка)")
        ws.cell(r, 4, student.full_name)
        r += 1
        put("(фамилия, имя, отчество)")
        put(f"{s.work_type or 'дипломный проект'} выполнен(а) на тему:")
        put(student.thesis_topic or "", align=LEFT)
        r += 1
        ws.cell(r, 1, "Под руководством")
        ws.cell(r, 3, student.supervisor or "")
        r += 2

        for line in [
            "В ГЭК представлены следующие документы:",
            "1. Справка об успеваемости студента;",
            "2. Дипломный проект;",
            "3. Демонстрационный материал;",
            "4. Отзыв руководителя;",
            "5. Рецензия",
        ]:
            ws.cell(r, 1, line)
            r += 1
        r += 1

        ws.cell(r, 1, "Студенту были заданы следующие вопросы:")
        r += 1
        questions = getattr(student, "questions", []) or []
        if questions:
            for qi, q in enumerate(questions, 1):
                asked_by = (q.asked_by or "").strip()
                prefix = f"{qi}. " + (f"{asked_by}: " if asked_by else "")
                ws.cell(r, 1, prefix + (q.question or ""))
                r += 1
        else:
            ws.cell(r, 1, "1. ____________________________________________")
            r += 1
        r += 1

        ws.cell(r, 1, "РЕШЕНИЕ ГЭК:").font = BOLD
        r += 1
        ws.cell(
            r, 1,
            f"1. Признать, что студент(ка) {student.full_name} "
            f"выполнил(а) и защитил(а) {s.work_type or 'дипломный проект'} "
            f"с оценкой «{student.grade or '___'}»"
        )
        r += 1
        ws.cell(
            r, 1,
            f"2. Присвоить квалификацию {student.qualification or ''} "
            f"по специальности {spec}"
        )
        r += 1
        honors = "с отличием" if student.with_honors else ""
        ws.cell(r, 1, f"3. Выдать диплом {honors}".rstrip())
        r += 2

        if chair:
            ws.cell(r, 1, "Председатель ГЭК")
            ws.cell(r, 7, chair[0].short_name or chair[0].full_name)
            r += 1
        for m in members:
            ws.cell(r, 1, "Член ГЭК")
            ws.cell(r, 7, m.short_name or m.full_name)
            r += 1
        if secretary:
            ws.cell(r, 1, "Секретарь ГЭК")
            ws.cell(r, 7, secretary[0].short_name or secretary[0].full_name)
            r += 1

        ws.column_dimensions["A"].width = 28
        for col in "BCDEFGH":
            ws.column_dimensions[col].width = 14